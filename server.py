"""
שרת Webhook רב-ערוצי (omnichannel) שמקבל הודעות נכנסות ומעדכן את ה-CRM.
בשלב זה השרת רץ מקומית; חיבור אמיתי לאינטרנט (ngrok) הוא שלב נפרד.

תמיכה בערוצים כרגע:
- whatsapp: פורמט Twilio האמיתי (form-encoded, שדות From/Body) - עובד בפועל, כולל
  מענה אוטומטי חזרה ללקוח (ראו process_message_with_reply ב-extract.py). זו תגובה
  בתוך חלון השיחה שהלקוח פתח, לא הודעה יזומה, ולכן אינה דורשת תבנית מאושרת מול מטא.
- instagram / facebook: פורמט JSON גנרי (source/contact_id/text) - PLACEHOLDER בלבד.
  חיבור אמיתי ל-Instagram/Facebook דורש אפליקציית Meta מאושרת ופענוח הפורמט
  האמיתי של Meta Graph API webhooks (שונה לגמרי מ-Twilio) - עוד לא ממומש כאן. אין
  עדיין מענה אוטומטי לערוצים אלה.

בהרצה ישירה (python server.py) השרת גם מפעיל את scheduler.py (Background Scheduler
Worker) כ-thread רקע שסורק תקופתית לידים קרים לפי חוקיות ורטיקל (contacts.csv) ומריץ
עליהם reactivate.py. ⚠️ ברירת המחדל היא dry-run בלבד - ראו האזהרה המפורטת בראש
scheduler.py לפני שמדליקים שליחה אוטומטית אמיתית (SCHEDULER_AUTO_SEND=true).
בפריסת production דרך Gunicorn (ראו Procfile) ה-thread הזה לא רץ בכלל (Gunicorn לא
מפעיל את if __name__=="__main__") - במקומו, scheduler_worker.py רץ כ-process נפרד
לגמרי (`clock`), כדי למנוע הכפלה בין כמה workers של Gunicorn.

--- הקשחה ל-production (2026-08-25) ---
- **משתני סביבה:** PORT/HOST/FLASK_DEBUG/LOG_LEVEL/VERIFY_TWILIO_SIGNATURE נקראים
  מ-.env עם ברירות מחדל בטוחות (ראו הבלוק "הגדרות מ-.env" למטה). FLASK_DEBUG חייב
  להיות false/לא-מוגדר ב-production אמיתי - מצב debug של Flask חושף Werkzeug
  debugger שמאפשר הרצת קוד שרירותי למי שמגיע לשגיאה לא מטופלת.
- **לוגים:** logging סטנדרטי של Python - קונסולה + קובץ מתגלגל (RotatingFileHandler)
  ל-server_error.log, בנפרד מ-chat_history.txt (שנשאר "לוג עסקי" של הודעות, לא לוג
  שגיאות טכני).
- **אימות Twilio:** כל בקשה ל-/webhook בפורמט Twilio (form-encoded, לא JSON) מאומתת
  מול חתימת X-Twilio-Signature (ראו _verify_twilio_request) - בלי זה, כל אחד שמנחש
  את כתובת ה-webhook יכול לשלוח בקשות מזויפות שיתפרשו כהודעות לקוח אמיתיות. השרת
  רץ מאחורי ngrok/reverse-proxy, ולכן יש ProxyFix כדי ש-request.url ישקף את הכתובת
  הציבורית האמיתית שטוויליו חתם עליה (אחרת האימות תמיד ייכשל בטעות).
"""

import csv
import io
import logging
import os
import sys
from datetime import datetime, timezone
from logging.handlers import RotatingFileHandler
from pathlib import Path
from xml.sax.saxutils import escape

from dotenv import load_dotenv
from flask import Flask, Response, jsonify, request, send_from_directory
from openpyxl import load_workbook
from twilio.request_validator import RequestValidator
from twilio.twiml.voice_response import Dial, VoiceResponse
from werkzeug.middleware.proxy_fix import ProxyFix

load_dotenv(dotenv_path=Path(__file__).parent / ".env")  # נתיב מפורש - עמיד לכל דרך הרצה/פריסה
# חייב לרוץ לפני ה-import-ים הבאים - db/reactivate/scheduler/extract קוראים os.environ בזמן טעינה

try:
    # כל הייבואים האלה - כולל db/reactivate/scheduler, לא רק extract/whatsapp_send
    # ישירות - תלויים ב-extract.py, שדורש ANTHROPIC_API_KEY בזמן טעינה (os.environ[...]).
    # לכן כל השרשרת חייבת להיות בתוך אותו try/except - אחרת ה-KeyError קורה כבר
    # ב-"import reactivate" למשל, לפני שמגיעים בכלל לבלוק שתופס אותו.
    import db
    import reactivate
    import scheduler
    import prompts
    import transcription
    import voice_call
    from extract import (
        DEFAULT_TENANT_ID,
        generate_call_summary,
        import_lead,
        load_customers,
        log_call_summary,
        log_manual_reply,
        process_message,
        process_message_with_reply,
        resolve_existing_phone,
        update_lead_agent,
        update_lead_category,
        update_lead_status,
        update_lead_voice_extraction,
    )
    from whatsapp_send import TWILIO_AUTH_TOKEN, _to_e164, is_trial_restriction, send_whatsapp_message
except KeyError as exc:
    # משתנה סביבה קריטי חסר (כרגע רק ANTHROPIC_API_KEY נדרש קשיח - os.environ[...] ולא
    # os.environ.get(...)) - נכשלים מיד עם הודעה ברורה, לא עם traceback גולמי שקשה
    # להבין ממנו מה בדיוק חסר. ב-Render: זה בדיוק מה שקורה אם שוכחים להגדיר משתנה
    # סביבה בדשבורד לפני ה-deploy הראשון - "Exit status 1" בלוג הוא הסימפטום החיצוני.
    print(f"שגיאת הגדרה: משתנה סביבה חסר - {exc}. בדקו את משתני הסביבה (.env מקומית / Render Environment).", file=sys.stderr)
    sys.exit(1)

# ---- הגדרות מ-.env (עם ברירות מחדל בטוחות) ----
PORT = int(os.environ.get("PORT", "5000"))
HOST = os.environ.get("HOST", "127.0.0.1")  # production מאחורי container/load-balancer: HOST=0.0.0.0
FLASK_DEBUG = os.environ.get("FLASK_DEBUG", "false").strip().lower() == "true"
LOG_LEVEL = os.environ.get("LOG_LEVEL", "INFO").upper()
VERIFY_TWILIO_SIGNATURE = os.environ.get("VERIFY_TWILIO_SIGNATURE", "true").strip().lower() == "true"

BASE_DIR = Path(__file__).parent
from paths import DATA_DIR  # noqa: E402 - chat_history.txt (state) נשמר כאן; server_error.log/index.html נשארים ב-BASE_DIR

# ---- לוגים: קונסולה + קובץ מתגלגל (לא גדל לאינסוף) ----
logger = logging.getLogger("whatsapp_crm")
logger.setLevel(LOG_LEVEL)
_log_formatter = logging.Formatter("%(asctime)s %(levelname)s [%(name)s] %(message)s")

_console_handler = logging.StreamHandler()
_console_handler.setFormatter(_log_formatter)
logger.addHandler(_console_handler)

_file_handler = RotatingFileHandler(
    BASE_DIR / "server_error.log", maxBytes=2_000_000, backupCount=5, encoding="utf-8"
)
_file_handler.setLevel(logging.WARNING)  # קובץ הלוג מתמקד בשגיאות/אזהרות - לא רעש כללי
_file_handler.setFormatter(_log_formatter)
logger.addHandler(_file_handler)

app = Flask(__name__)
# מאחורי ngrok/reverse-proxy: משקף את ה-scheme/host/port הציבוריים האמיתיים מתוך
# X-Forwarded-* במקום 127.0.0.1 המקומי - קריטי גם לאימות חתימת Twilio (_verify_twilio_request)
app.wsgi_app = ProxyFix(app.wsgi_app, x_for=1, x_proto=1, x_host=1, x_port=1)

_twilio_validator = RequestValidator(TWILIO_AUTH_TOKEN) if TWILIO_AUTH_TOKEN else None


def _verify_twilio_request(req) -> bool:
    """מאמת שבקשת ה-webhook הגיעה באמת מ-Twilio, לפי חתימת X-Twilio-Signature (HMAC
    עם ה-Auth Token) - https://www.twilio.com/docs/usage/webhooks/webhooks-security.
    בלי זה, כל אחד שמנחש את כתובת ה-webhook יכול לשלוח בקשות מזויפות."""
    if _twilio_validator is None:
        logger.warning("אימות Twilio מבוקש אבל TWILIO_AUTH_TOKEN לא מוגדר ב-.env - דוחה את הבקשה")
        return False
    signature = req.headers.get("X-Twilio-Signature", "")
    return _twilio_validator.validate(req.url, req.form, signature)


@app.errorhandler(Exception)
def handle_uncaught_exception(exc):
    logger.error("שגיאה לא מטופלת בבקשה ל-%s: %s", request.path, exc, exc_info=True)
    return jsonify({"error": "שגיאת שרת פנימית"}), 500


SUPPORTED_SOURCES = {"whatsapp", "instagram", "facebook"}
CHAT_HISTORY_FILE = DATA_DIR / "chat_history.txt"


def parse_incoming() -> tuple[str, str, str]:
    """מנרמל הודעה נכנסת מכל ערוץ לפורמט אחיד: (contact_id, message_text, source)."""
    source = request.args.get("source", "whatsapp")
    if source not in SUPPORTED_SOURCES:
        source = "whatsapp"

    if request.is_json:
        # PLACEHOLDER: פורמט גנרי, לא הפורמט האמיתי של Meta Graph API
        data = request.get_json(silent=True) or {}
        source = data.get("source", source)
        contact_id = str(data.get("contact_id", ""))
        message_text = data.get("text", "")
    else:
        # Twilio (WhatsApp) שולח form-encoded עם השדות From ו-Body
        contact_id = request.form.get("From", "").replace("whatsapp:", "")
        message_text = request.form.get("Body", "")

    return contact_id, message_text, source


def _log_incoming_message(tenant_id: str, source: str, contact_id: str, message_text: str) -> None:
    """מוסיף רשומה מסודרת של הודעה נכנסת מליד ל-chat_history.txt (לוג טקסטואלי,
    בנוסף לרישום המובנה בהיסטוריית הכרטיס בתוך customers.json)."""
    timestamp = datetime.now(timezone.utc).isoformat()
    line = f"[{timestamp}] tenant={tenant_id} source={source} from={contact_id}: {message_text}\n"
    with CHAT_HISTORY_FILE.open("a", encoding="utf-8") as f:
        f.write(line)


@app.route("/")
def index():
    return send_from_directory(BASE_DIR, "index.html")


@app.route("/api/leads")
def api_leads():
    """רשימת כל הלידים מ-customers.json (מקור האמת) - לשימוש הדשבורד ב-index.html.
    ?include_last_message=1 (משמש את תצוגת ה-Unified Inbox): מוסיף לכל ליד את ההודעה
    האחרונה מטבלת messages (תצוגה מקדימה + מיון לפי פעילות) - כבוי כברירת מחדל כדי
    לא להאט את טעינת הטבלה הרגילה עם שאילתת DB נוספת לכל שורה.
    ?include_score=1 (משמש את הטבלה הראשית + מסנן טווח הציון): מוסיף "ציון חום"
    מחושב (1-100, ראו db.compute_lead_score) - גם הוא כבוי כברירת מחדל מאותה סיבה
    (עוד כמה שאילתות DB לכל שורה)."""
    include_last_message = request.args.get("include_last_message") == "1"
    include_score = request.args.get("include_score") == "1"

    leads = []
    for key, card in load_customers().items():
        tenant_id, _, key_phone = key.partition("::")
        phone = card.get("phone", key_phone)
        resolved_tenant_id = card.get("tenant_id", tenant_id)
        lead = {
            "tenant_id": resolved_tenant_id,
            "phone": phone,
            "customer_name": card.get("customer_name"),
            "business_name": card.get("business_name"),
            "location": card.get("location"),
            "source_channel": card.get("source_channel"),
            "import_source": card.get("import_source"),
            "lead_status": card.get("lead_status"),
            "category": card.get("category"),
            "agent": card.get("agent"),
            "property_type": card.get("property_type"),
            "budget": card.get("budget"),
        }
        if include_last_message:
            last = db.get_last_message(phone, tenant_id=resolved_tenant_id)
            lead["last_message"] = last["message"] if last else None
            lead["last_message_at"] = last["timestamp"] if last else None
            lead["last_message_direction"] = last["direction"] if last else None
        if include_score:
            lead["score"] = db.compute_lead_score(phone, tenant_id=resolved_tenant_id)
        leads.append(lead)
    return jsonify(leads)


@app.route("/api/leads", methods=["POST"])
def api_create_lead():
    """הוספת ליד בודד ידנית מהדשבורד ("➕ הוסף ליד") - בניגוד ל-/api/leads/import
    (קובץ שלם), כאן שורה אחת. עוטף את import_lead בדיוק כמו הייבוא - אותה לוגיקת
    upsert, בלי כפילות קוד. הטלפון עובר resolve_existing_phone (לא _to_e164 ישירות)
    כדי לא ליצור כרטיס כפול אם הליד כבר קיים בפורמט מקומי - ראו התיעוד שם."""
    data = request.get_json(silent=True) or {}
    phone = (data.get("phone") or "").strip()
    if not phone:
        return jsonify({"error": "חסר טלפון"}), 400

    tenant_id = data.get("tenant_id") or DEFAULT_TENANT_ID
    status = data.get("lead_status")
    if status not in VALID_LEAD_STATUSES:
        status = None

    card, is_new = import_lead(
        resolve_existing_phone(phone, tenant_id=tenant_id),
        tenant_id=tenant_id,
        customer_name=(data.get("customer_name") or "").strip() or None,
        business_name=(data.get("business_name") or "").strip() or None,
        location=(data.get("location") or "").strip() or None,
        lead_status=status,
        category=(data.get("category") or "").strip() or None,
        agent=(data.get("agent") or "").strip() or None,
    )
    return jsonify({"ok": True, "card": card, "is_new": is_new})


@app.route("/api/search")
def api_search():
    """חיפוש חופשי רב-שדות: קודם בודק התאמה ישירה בשדות הכרטיס (שם/עסק/מיקום/טלפון/
    קטגוריה, מ-customers.json - בזיכרון, זול), ובנוסף מחפש בתוכן פעילות (הודעות
    וואטסאפ, הערות/תקציר שיחות, כותרת/הערות משימות - db.search_activity). מחזיר
    איחוד (union) של שתי ההתאמות כרשימת {phone, tenant_id} - הדשבורד מסנן לפיה את
    allLeads הטעון כבר, בלי לשלוף מחדש את כל רשימת הלידים."""
    q = (request.args.get("q") or "").strip()
    if not q:
        return jsonify([])
    q_lower = q.lower()

    matches = set()
    for key, card in load_customers().items():
        tenant_id, _, key_phone = key.partition("::")
        phone = card.get("phone", key_phone)
        resolved_tenant_id = card.get("tenant_id", tenant_id)
        haystack = " ".join(str(card.get(f) or "") for f in
                             ("customer_name", "business_name", "location", "phone", "category", "agent")).lower()
        if q_lower in haystack:
            matches.add((phone, resolved_tenant_id))

    matches.update(db.search_activity(q))

    return jsonify([{"phone": phone, "tenant_id": tenant_id} for phone, tenant_id in matches])


VALID_LEAD_STATUSES = {"new", "contacted", "hot", "not_relevant"}


@app.route("/api/leads/status", methods=["POST"])
def api_update_lead_status():
    """מעדכן lead_status ידנית מהדשבורד (שורת הטבלה או פאנל ההיסטוריה). זו רק עדכון
    סטטוס ב-customers.json - לא הודעה, ולכן לא נרשם בטבלת messages."""
    data = request.get_json(silent=True) or {}
    phone = (data.get("phone") or "").strip()
    tenant_id = data.get("tenant_id") or DEFAULT_TENANT_ID
    status = data.get("status")

    if not phone or status not in VALID_LEAD_STATUSES:
        return jsonify({"error": "טלפון או סטטוס לא תקינים"}), 400

    card = update_lead_status(phone, status, tenant_id=tenant_id)
    return jsonify({"ok": True, "card": card})


@app.route("/api/leads/category", methods=["POST"])
def api_update_lead_category():
    """מעדכן קטגוריה חופשית לליד (לדוגמה: נדל"ן/פרטי/משפחה) - שדה סיווג ידני,
    לא קשור ל-lead_status. category ריק ("") מנקה את השדה."""
    data = request.get_json(silent=True) or {}
    phone = (data.get("phone") or "").strip()
    tenant_id = data.get("tenant_id") or DEFAULT_TENANT_ID
    category = (data.get("category") or "").strip()

    if not phone:
        return jsonify({"error": "חסר טלפון"}), 400

    card = update_lead_category(phone, category, tenant_id=tenant_id)
    return jsonify({"ok": True, "card": card})


@app.route("/api/leads/agent", methods=["POST"])
def api_update_lead_agent():
    """מעדכן "סוכן מטפל" - שדה טקסט חופשי, מטא-דאטה בלבד (כמו /api/leads/category)."""
    data = request.get_json(silent=True) or {}
    phone = (data.get("phone") or "").strip()
    tenant_id = data.get("tenant_id") or DEFAULT_TENANT_ID
    agent = (data.get("agent") or "").strip()

    if not phone:
        return jsonify({"error": "חסר טלפון"}), 400

    card = update_lead_agent(phone, agent, tenant_id=tenant_id)
    return jsonify({"ok": True, "card": card})


# ייבוא לידים מ-CSV/Excel - מיפוי אוטומטי של עמודות (בעברית או באנגלית) לשדות הסטנדרטיים
ALLOWED_IMPORT_EXTENSIONS = {".csv", ".xlsx"}
IMPORT_FIELD_ALIASES = {
    "customer_name": {"name", "customer_name", "full name", "שם", "שם לקוח", "שם מלא"},
    "phone": {"phone", "phone_number", "mobile", "טלפון", "מספר טלפון", "נייד", "מס' טלפון"},
    "business_name": {"business", "business_name", "company", "עסק", "שם עסק", "חברה"},
    "location": {"location", "city", "מיקום", "עיר", "אזור"},
    "import_source": {"source", "lead_source", "מקור", "מקור ליד", "ערוץ מקור"},
    "lead_status": {"status", "lead_status", "סטטוס"},
    "category": {"category", "vertical", "קטגוריה", "סיווג"},
    "agent": {"agent", "assigned_agent", "owner", "סוכן", "סוכן מטפל", "נציג"},
}


def _normalize_header(h) -> str:
    return str(h or "").strip().lower()


def _map_import_row(raw_row: dict) -> dict:
    """ממפה שורת CSV/Excel גולמית (עמודות בכל שם סביר, עברית או אנגלית) לשדות התקניים."""
    normalized = {_normalize_header(k): v for k, v in raw_row.items()}
    mapped = {}
    for field, aliases in IMPORT_FIELD_ALIASES.items():
        for alias in aliases:
            value = normalized.get(_normalize_header(alias))
            if value not in (None, ""):
                mapped[field] = str(value).strip()
                break
    return mapped


def _parse_import_csv(file_stream) -> list[dict]:
    text = file_stream.read().decode("utf-8-sig")  # utf-8-sig סופג BOM מ-CSV שיוצא מ-Excel
    return list(csv.DictReader(io.StringIO(text)))


def _parse_import_excel(file_stream) -> list[dict]:
    workbook = load_workbook(file_stream, read_only=True, data_only=True)
    sheet = workbook.active
    rows_iter = sheet.iter_rows(values_only=True)
    try:
        headers = [str(h).strip() if h is not None else "" for h in next(rows_iter)]
    except StopIteration:
        return []
    rows = []
    for row in rows_iter:
        if all(cell is None for cell in row):
            continue
        rows.append({headers[i]: row[i] for i in range(min(len(headers), len(row)))})
    return rows


@app.route("/api/leads/import", methods=["POST"])
def api_import_leads():
    """מייבא לידים מקובץ CSV/XLSX שהועלה מהדשבורד. ממפה אוטומטית שם/טלפון/עסק/מקור/
    סטטוס (בעברית או באנגלית), ומונע כפילויות ע"י upsert לפי מספר טלפון מנורמל
    (E.164) + tenant_id - ראו import_lead ב-extract.py."""
    if "file" not in request.files:
        return jsonify({"error": "לא צורף קובץ (שדה 'file')"}), 400

    upload = request.files["file"]
    ext = Path(upload.filename or "").suffix.lower()
    if ext not in ALLOWED_IMPORT_EXTENSIONS:
        return jsonify({"error": f"סוג קובץ לא נתמך: '{ext or 'ללא סיומת'}'. נתמכים: CSV, XLSX"}), 400

    tenant_id = request.form.get("tenant_id") or DEFAULT_TENANT_ID

    try:
        raw_rows = _parse_import_csv(upload.stream) if ext == ".csv" else _parse_import_excel(upload.stream)
    except Exception as exc:
        return jsonify({"error": f"שגיאה בקריאת הקובץ: {exc}"}), 400

    imported = updated = 0
    skipped = []
    for row_num, raw_row in enumerate(raw_rows, start=2):  # שורה 1 = כותרות
        mapped = _map_import_row(raw_row)
        phone = mapped.get("phone")
        if not phone:
            skipped.append({"row": row_num, "reason": "אין מספר טלפון"})
            continue

        status = mapped.get("lead_status", "").lower()
        if status not in VALID_LEAD_STATUSES:
            status = None  # סטטוס לא מוכר - מתעלמים ממנו, לא נכשלים על כל השורה

        _card, is_new = import_lead(
            resolve_existing_phone(phone, tenant_id=tenant_id),
            tenant_id=tenant_id,
            customer_name=mapped.get("customer_name"),
            business_name=mapped.get("business_name"),
            location=mapped.get("location"),
            lead_status=status,
            import_source=mapped.get("import_source"),
            category=mapped.get("category"),
            agent=mapped.get("agent"),
        )
        if is_new:
            imported += 1
        else:
            updated += 1

    return jsonify({
        "ok": True,
        "total_rows": len(raw_rows),
        "imported": imported,
        "updated": updated,
        "skipped": skipped,
    })


EXPORT_COLUMNS = [
    "customer_name", "phone", "business_name", "location", "agent", "category",
    "lead_status", "score", "source_channel", "import_source", "tenant_id",
]


@app.route("/api/leads/export")
def api_export_leads():
    """מייצא את כל הלידים (כולל ציון חום מחושב) ל-CSV נקי - "📤 ייצוא לידים" בדשבורד.
    העמודות תואמות בכוונה למה שנתמך גם בייבוא חזרה (IMPORT_FIELD_ALIASES) - חוץ
    מ-score, שהוא שדה מחושב-נגזר (db.compute_lead_score) ולעולם לא נשמר/מיובא."""
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(EXPORT_COLUMNS)
    for key, card in load_customers().items():
        tenant_id, _, key_phone = key.partition("::")
        phone = card.get("phone", key_phone)
        resolved_tenant_id = card.get("tenant_id", tenant_id)
        score = db.compute_lead_score(phone, tenant_id=resolved_tenant_id)
        writer.writerow([
            card.get("customer_name") or "", phone, card.get("business_name") or "",
            card.get("location") or "", card.get("agent") or "", card.get("category") or "",
            card.get("lead_status") or "", score, card.get("source_channel") or "",
            card.get("import_source") or "", resolved_tenant_id,
        ])
    csv_bytes = output.getvalue().encode("utf-8-sig")  # BOM כדי ש-Excel יפתח עברית נכון
    return Response(
        csv_bytes, mimetype="text/csv",
        headers={"Content-Disposition": "attachment; filename=leads_export.csv"},
    )


@app.route("/api/messages")
def api_messages():
    """היסטוריית ההודעות של ליד מסוים מטבלת messages ב-crm_data.db - לשימוש הדשבורד.
    since (אופציונלי): מחזיר רק הודעות חדשות יותר - משמש את ה-polling של חלון הצ'אט
    החי (ראו index.html) כדי לרענן בלי לשלוף את כל ההיסטוריה בכל בדיקה."""
    phone = request.args.get("phone", "")
    tenant_id = request.args.get("tenant_id", DEFAULT_TENANT_ID)
    since = request.args.get("since") or None
    if not phone:
        return Response(status=400)
    return jsonify(db.get_messages(phone, tenant_id=tenant_id, since=since))


@app.route("/api/reactivate", methods=["POST"])
def api_reactivate():
    """מפעיל את קמפיין חימום הלידים הקרים (reactivate.py) מתוך הדשבורד. ברירת המחדל
    היא preview בלבד (send=False, כברירת המחדל של reactivate.py) - לא נשלח שום דבר.
    שליחה בפועל דורשת send:true מפורש בגוף הבקשה; זה האישור האנושי הנדרש (קליק מפורש
    בממשק אחרי צפייה בתצוגה המקדימה) - ראו מדיניות הבטיחות ב-CLAUDE.md.
    days (אופציונלי, ברירת מחדל DEFAULT_COLD_DAYS=30): סף "לא נוצר קשר מעל X ימים" -
    ראו reactivate.get_cold_leads למה זה לא ברירת המחדל הגלובלית של reactivate.py."""
    data = request.get_json(silent=True) or {}
    tenant_id = data.get("tenant_id") or DEFAULT_TENANT_ID
    send = bool(data.get("send", False))
    try:
        days = int(data.get("days", reactivate.DEFAULT_COLD_DAYS))
    except (TypeError, ValueError):
        days = reactivate.DEFAULT_COLD_DAYS

    results = reactivate.run_reactivation_campaign(tenant_id=tenant_id, send=send, days=days)
    return jsonify({"send": send, "days": days, "count": len(results), "results": results})


@app.route("/api/scheduler/status")
def api_scheduler_status():
    """מצב מנוע התזמון האוטומטי (scheduler.py) - קריאה בלבד. חשוב לבדוק את auto_send:
    כשהוא False (ברירת המחדל) המנוע רץ במצב dry-run בלבד ולא שולח הודעות אמיתיות."""
    return jsonify(scheduler.status())


@app.route("/api/scheduler/runs")
def api_scheduler_runs():
    """היסטוריית מחזורי הסריקה האחרונים של מנוע התזמון, מטבלת scheduler_runs."""
    return jsonify(db.get_scheduler_runs())


@app.route("/api/scheduler/run-now", methods=["POST"])
def api_scheduler_run_now():
    """מפעיל מחזור סריקה אחד מיידית, בלי לחכות למרווח התזמון (שימושי לבדיקה/דמו).
    מכבד את אותה הגדרת SCHEDULER_AUTO_SEND כמו הרצה רגילה - לא שולח הודעות אמיתיות
    אם auto_send=False."""
    data = request.get_json(silent=True) or {}
    tenant_id = data.get("tenant_id") or DEFAULT_TENANT_ID
    results = scheduler.run_scan(tenant_id=tenant_id)
    return jsonify({"auto_send": scheduler.AUTO_SEND, "count": len(results), "results": results})


@app.route("/api/messages/send", methods=["POST"])
def api_send_message():
    """שליחת הודעה ידנית בפועל דרך Twilio, מתוך הדשבורד. כל קריאה כאן מגיעה מקליק
    מפורש של נציג אנושי על "שלח" - זה האישור האנושי הנדרש למדיניות השליחה בפרויקט
    (ראו CLAUDE.md). שים לב: הודעה ליד שלא כתב הודעה ב-24 השעות האחרונות נחשבת
    "business-initiated conversation" ועלולה לדרוש הודעת-תבנית מאושרת מול מטא."""
    data = request.get_json(silent=True) or {}
    phone = (data.get("phone") or "").strip()
    tenant_id = data.get("tenant_id") or DEFAULT_TENANT_ID
    message = (data.get("message") or "").strip()

    if not phone or not message:
        return jsonify({"error": "חסר טלפון או תוכן הודעה"}), 400

    simulated = False
    sid = None
    try:
        sid = send_whatsapp_message(phone, message)
    except Exception as exc:
        if not is_trial_restriction(exc):
            return jsonify({"error": str(exc)}), 502
        # חשבון Twilio מסוג Trial חסם את השליחה בפועל (למשל נמען לא מאומת) - זו לא
        # שגיאת קוד; רושמים את ההודעה כ"מדומה" (simulated) כדי לאפשר לבדוק את חלון
        # השיחות והזרימה בדשבורד בלי להיחסם. הלקוח לא קיבל את ההודעה בפועל.
        simulated = True

    card = log_manual_reply(phone, message, tenant_id=tenant_id, simulated=simulated)
    db.log_message(phone, message, direction="out", tenant_id=tenant_id, channel="whatsapp", simulated=simulated)

    response = {"ok": True, "sid": sid, "card": card, "simulated": simulated}
    if simulated:
        response["warning"] = (
            "חשבון Twilio מסוג Trial חסם את השליחה בפועל (נמען לא מאומת) - "
            "ההודעה נרשמה כסימולציה לצורך בדיקה, אך לא נשלחה ללקוח."
        )
    return jsonify(response)


@app.route("/api/calls/start", methods=["POST"])
def api_calls_start():
    """יוזם Click-to-Call (גישור נציג→לקוח). אישור אנושי = הקליק על '📞 שיחה' בדשבורד.
    כרגע (סביבת פיתוח) אין מספר Twilio Voice/PUBLIC_BASE_URL מוגדרים - זה ייתפס כאן
    ויתנהג בדיוק כמו is_trial_restriction ב-/api/messages/send: נרשם כ-simulated,
    מוחזר 200 (לא 502), כדי לאפשר להמשיך ולבדוק את שאר הזרימה (הערות → תקציר → כרטיס)."""
    data = request.get_json(silent=True) or {}
    phone = (data.get("phone") or "").strip()
    tenant_id = data.get("tenant_id") or DEFAULT_TENANT_ID
    if not phone:
        return jsonify({"error": "חסר טלפון"}), 400

    simulated, call_sid, status = False, None, "initiated"
    try:
        call_sid = voice_call.start_bridge_call(phone, tenant_id=tenant_id)
    except RuntimeError as exc:
        logger.warning("שיחת Voice לא הופעלה (תצורה חסרה): %s", exc)
        simulated, status = True, "simulated_no_config"
    except Exception as exc:
        if not is_trial_restriction(exc):
            return jsonify({"error": str(exc)}), 502
        simulated, status = True, "simulated_trial_restriction"

    call_id = db.create_call(phone, tenant_id=tenant_id, call_sid=call_sid, status=status, simulated=simulated)
    response = {"ok": True, "call_id": call_id, "call_sid": call_sid, "simulated": simulated, "status": status}
    if simulated:
        response["warning"] = (
            "שיחה אמיתית לא בוצעה (חסרה תצורת Voice ב-.env, או שחשבון Trial חוסם) - "
            "נרשמה כסימולציה כדי לאפשר להמשיך למילוי הערות ותקציר."
        )
    return jsonify(response)


@app.route("/api/calls/<int:call_id>/notes", methods=["POST"])
def api_call_notes(call_id):
    """הנציג שולח הערות חופשיות שהקליד אחרי השיחה → Claude מייצר תקציר
    (extract.generate_call_summary) → נשמר ב-calls (audit trail: הערות + תקציר) וגם
    משוקף להיסטוריה/messages (extract.log_call_summary + db.log_message, channel=
    "voice") - כך שהוא מופיע אוטומטית בצ'אט הקיים (פאנל/Inbox) דרך אותו polling
    שכבר קיים, בלי קוד רינדור נוסף. עובד גם על call_id שנוצר במצב simulated - ולכן
    ניתן לבדיקה מלאה גם בלי שיחת Twilio אמיתית."""
    data = request.get_json(silent=True) or {}
    notes = (data.get("notes") or "").strip()
    if not notes:
        return jsonify({"error": "חסרות הערות"}), 400

    call = db.get_call(call_id)
    if not call:
        return jsonify({"error": "שיחה לא נמצאה"}), 404

    phone, tenant_id, simulated = call["phone"], call["tenant_id"], call["simulated"]
    card = load_customers().get(f"{tenant_id}::{phone}", {})
    summary = generate_call_summary(notes, card)

    updated_call = db.save_call_notes_and_summary(call_id, notes, summary)
    log_call_summary(phone, summary, tenant_id=tenant_id, simulated=simulated)
    db.log_message(phone, summary, direction="out", tenant_id=tenant_id, channel="voice", simulated=simulated)

    return jsonify({"ok": True, "call": updated_call, "summary": summary})


@app.route("/api/calls/<int:call_id>/transcribe", methods=["POST"])
def api_call_transcribe(call_id):
    """מתמלל קובץ אודיו שהועלה (transcription.transcribe_audio, OpenAI Whisper API)
    ומחזיר את הטקסט - **לא** שומר כלום בעצמו. ה-UI ממלא את טקסט התמלול לתוך אותה
    תיבת הערות שממנה ייצור התקציר (POST /api/calls/<id>/notes) - כך שהתמלול הוא
    רק דרך חלופית למלא את ההערות, וכל שאר הצינור (שמירה בכרטיס + חותמת זמן +
    היסטוריה) זהה לגמרי לזרימת ההקלדה הידנית הקיימת, בלי קוד כפול. אם
    OPENAI_API_KEY לא מוגדר - מחזיר 400 עם הודעה ברורה, וה-UI חוזר להקלדה ידנית."""
    if not db.get_call(call_id):
        return jsonify({"error": "שיחה לא נמצאה"}), 404

    if "audio" not in request.files:
        return jsonify({"error": "לא צורף קובץ אודיו (שדה 'audio')"}), 400

    upload = request.files["audio"]
    try:
        text = transcription.transcribe_audio(upload.read(), upload.filename or "recording.webm")
    except RuntimeError as exc:
        return jsonify({"error": str(exc)}), 400
    except Exception as exc:
        logger.error("תמלול אודיו נכשל (call_id=%s): %s", call_id, exc, exc_info=True)
        return jsonify({"error": f"תמלול נכשל: {exc}"}), 502

    return jsonify({"ok": True, "text": text})


VALID_PROMPT_VERTICALS = {"ecommerce", "services", "real_estate"}


@app.route("/api/prompts")
def api_get_prompts():
    """כל תבניות/פרומפטים הסוכנים (Speed-to-Lead, החייאה מותאם-ענף, תיאום פולו-אפ,
    חילוץ פרטים) - לתצוגה/עריכה מה-CRM ("🤖 תבניות וסוכנים"). ראו prompts.py."""
    return jsonify(prompts.get_all_prompts())


@app.route("/api/prompts/<key>", methods=["POST"])
def api_update_prompt(key):
    """מעדכן תבנית פרומפט. vertical (אופציונלי, רק ל-reactivation_outreach -
    ecommerce/services/real_estate): עורך override מותאם-ענף במקום התבנית
    הבסיסית - template ריק עם vertical מוחק את ה-override (חזרה לברירת המחדל)."""
    data = request.get_json(silent=True) or {}
    template = (data.get("template") or "").strip()
    vertical = data.get("vertical") or None
    if vertical and vertical not in VALID_PROMPT_VERTICALS:
        return jsonify({"error": "ורטיקל לא תקין"}), 400
    if not template and not vertical:
        return jsonify({"error": "חסרה תבנית"}), 400

    try:
        entry = prompts.update_prompt(key, template, vertical=vertical)
    except KeyError:
        return jsonify({"error": "תבנית לא נמצאה"}), 404

    return jsonify({"ok": True, "prompt": entry})


@app.route("/api/calls")
def api_calls():
    """יומן שיחות Voice של ליד - בדומה ל-/api/messages, אבל לטבלת calls (שיחה = שורה,
    לא הודעה בודדת)."""
    phone = request.args.get("phone", "")
    tenant_id = request.args.get("tenant_id", DEFAULT_TENANT_ID)
    if not phone:
        return Response(status=400)
    return jsonify(db.get_calls(phone, tenant_id=tenant_id))


VALID_TASK_STATUSES = {"pending", "done", "cancelled"}


@app.route("/api/tasks", methods=["GET", "POST"])
def api_tasks():
    """משימות מעקב/תזכורות (follow-up) - לא קשור להודעות. GET עם ?phone=&tenant_id=
    מחזיר את המשימות של ליד ספציפי (לפאנל "📅 משימות"); GET בלי phone מחזיר את כל
    המשימות (אופציונלית מסונן ב-?tenant_id=&status=, לתצוגת "📅 יומן" הגלובלית).
    POST יוצר משימה חדשה."""
    if request.method == "GET":
        phone = request.args.get("phone") or None
        tenant_id = request.args.get("tenant_id") or None
        status = request.args.get("status") or None
        return jsonify(db.get_tasks(phone=phone, tenant_id=tenant_id, status=status))

    data = request.get_json(silent=True) or {}
    phone = (data.get("phone") or "").strip()
    tenant_id = data.get("tenant_id") or DEFAULT_TENANT_ID
    title = (data.get("title") or "").strip()
    due_date = (data.get("due_date") or "").strip()
    due_time = (data.get("due_time") or "").strip() or None
    notes = (data.get("notes") or "").strip() or None

    if not phone or not title or not due_date:
        return jsonify({"error": "חסר טלפון, כותרת או תאריך יעד"}), 400

    task_id = db.create_task(phone, tenant_id=tenant_id, title=title, due_date=due_date, due_time=due_time, notes=notes)
    return jsonify({"ok": True, "task_id": task_id})


@app.route("/api/tasks/<int:task_id>/status", methods=["POST"])
def api_task_status(task_id):
    """מעדכן סטטוס משימה (pending/done/cancelled) - מהתגית "✓ בוצע"/"✗ בטל" בפאנל
    המשימות או ביומן הגלובלי."""
    data = request.get_json(silent=True) or {}
    status = data.get("status")
    if status not in VALID_TASK_STATUSES:
        return jsonify({"error": "סטטוס לא תקין"}), 400

    task = db.update_task_status(task_id, status)
    if not task:
        return jsonify({"error": "משימה לא נמצאה"}), 404
    return jsonify({"ok": True, "task": task})


VALID_FEEDBACK_TYPES = {"bug", "idea"}
VALID_FEEDBACK_STATUSES = {"new", "in_progress", "done", "wontfix"}


@app.route("/api/feedback", methods=["GET", "POST"])
def api_feedback():
    """משוב על המערכת עצמה (באג/רעיון) - לא קשור ללידים בכלל, פנימי לצוות שמפעיל
    את ה-CRM. GET מחזיר את כל הרשומות (החדש ביותר קודם); POST יוצר רשומה חדשה."""
    if request.method == "GET":
        return jsonify(db.get_feedback())

    data = request.get_json(silent=True) or {}
    feedback_type = data.get("feedback_type")
    description = (data.get("description") or "").strip()
    if feedback_type not in VALID_FEEDBACK_TYPES or not description:
        return jsonify({"error": "סוג משוב או תיאור לא תקינים"}), 400

    feedback_id = db.create_feedback(feedback_type, description)
    return jsonify({"ok": True, "feedback_id": feedback_id})


@app.route("/api/feedback/<int:feedback_id>/status", methods=["POST"])
def api_feedback_status(feedback_id):
    """מעדכן סטטוס טיפול במשוב (new/in_progress/done/wontfix)."""
    data = request.get_json(silent=True) or {}
    status = data.get("status")
    if status not in VALID_FEEDBACK_STATUSES:
        return jsonify({"error": "סטטוס לא תקין"}), 400

    feedback = db.update_feedback_status(feedback_id, status)
    if not feedback:
        return jsonify({"error": "משוב לא נמצא"}), 404
    return jsonify({"ok": True, "feedback": feedback})


@app.route("/voice/connect", methods=["POST"])
def voice_connect():
    """TwiML: Twilio מגיע לכאן ברגע שהנציג (הרגל הראשונה של הגישור) עונה. מגשר
    (<Dial>) למספר הלקוח, עם הקלטה (record="record-from-answer") ו-callback כשההקלטה
    מוכנה. אין תמלול אוטומטי - הוחלט במפורש (ראו voice_call.py)."""
    if VERIFY_TWILIO_SIGNATURE and not _verify_twilio_request(request):
        logger.warning("בקשת /voice/connect נדחתה - חתימת Twilio לא תקינה/חסרה")
        return Response(status=403)

    customer_phone = request.args.get("customer_phone", "")
    if not customer_phone:
        return Response(status=400)

    response = VoiceResponse()
    dial = Dial(
        record="record-from-answer",
        recording_status_callback=f"{voice_call.PUBLIC_BASE_URL}/voice/recording-status",
        recording_status_callback_event=["completed"],
    )
    dial.number(customer_phone)
    response.append(dial)
    return Response(str(response), mimetype="text/xml")


@app.route("/voice/status", methods=["POST"])
def voice_status():
    """Status callback לרגל הראשונה של הגשר (הנציג) - CallSid/CallStatus/CallDuration."""
    if VERIFY_TWILIO_SIGNATURE and not _verify_twilio_request(request):
        return Response(status=403)
    call_sid = request.form.get("CallSid", "")
    status = request.form.get("CallStatus", "")
    duration = request.form.get("CallDuration")
    if call_sid and status:
        db.update_call_status(call_sid, status, duration_seconds=int(duration) if duration else None)
    return Response(status=204)


@app.route("/voice/recording-status", methods=["POST"])
def voice_recording_status():
    """הקלטת השיחה המגושרת מוכנה - שומר רק RecordingUrl. אין תמלול/STT אוטומטי
    (הוחלט במפורש - ראו "Recording + הערות ידניות" ב-CLAUDE.md)."""
    if VERIFY_TWILIO_SIGNATURE and not _verify_twilio_request(request):
        return Response(status=403)
    call_sid = request.form.get("CallSid", "")
    recording_url = request.form.get("RecordingUrl", "")
    if call_sid and recording_url:
        db.save_recording_url(call_sid, recording_url)
    return Response(status=204)


@app.route("/webhook", methods=["POST"])
@app.route("/webhook/<tenant_id>", methods=["POST"])
def webhook(tenant_id: str = DEFAULT_TENANT_ID):
    # כל עסק (tenant) מקבל כתובת webhook משלו עם ה-tenant_id שלו בנתיב, למשל:
    # https://<ngrok-url>/webhook/business_a - כך שהנתונים של כל עסק מבודדים זה מזה.

    # אימות Twilio חל רק על בקשות בפורמט Twilio האמיתי (form-encoded) - לא על
    # ה-JSON הגנרי של instagram/facebook (PLACEHOLDER, לא תעבורת Twilio אמיתית).
    if VERIFY_TWILIO_SIGNATURE and not request.is_json:
        if not _verify_twilio_request(request):
            logger.warning(
                "בקשת webhook נדחתה - חתימת X-Twilio-Signature לא תקינה/חסרה (מ-%s, tenant=%s)",
                request.remote_addr, tenant_id,
            )
            return Response(status=403)

    contact_id, message_text, source = parse_incoming()

    # הודעה קולית נכנסת (WhatsApp voice note): Twilio שולח Body ריק ורק מדיה
    # (NumMedia/MediaUrl0/MediaContentType0) - בלי הטיפול הזה message_text היה
    # נשאר ריק וההודעה כולה נדחית ב-400 למטה, בלי להירשם בשום מקום ("נעלמת"
    # מה-Inbox). מתמלל אוטומטית (transcription.transcribe_incoming_voice_message,
    # OpenAI Whisper) ומשתמש בטקסט המתומלל בדיוק כמו הודעת טקסט רגילה מכאן והלאה -
    # אותו צינור process_message_with_reply/db.log_message, בלי קוד כפול. אם
    # התמלול עצמו נכשל (OPENAI_API_KEY לא מוגדר/לא תקין, שגיאת רשת/API) - עדיין
    # מקבלים טקסט placeholder ברור (⚠️) במקום None, כדי שההודעה עדיין תירשם
    # ותופיע ב-Inbox עם סימון שקרה כשל, לא תיעלם בשקט.
    if not message_text and source == "whatsapp" and not request.is_json:
        voice_result = transcription.transcribe_incoming_voice_message(request.form)
        if voice_result:
            if voice_result["success"]:
                message_text = f"🎙️ {voice_result['text']}"
                # חילוץ שדות אוטומטי (שם/סוג נכס/תקציב - ממוקד נדל"ן, ראו "חזון
                # המוצר" ב-CLAUDE.md) מהתמלול, עם gpt-4o-mini - רק אחרי תמלול
                # מוצלח (אין טעם לנתח טקסט placeholder של כישלון). כשל בשלב הזה
                # לא אמור להפיל את כל הטיפול בהודעה - היא כבר תירשם כרגיל בהמשך
                # גם אם חילוץ השדות נכשל.
                if contact_id:
                    try:
                        fields = transcription.extract_voice_message_fields(voice_result["text"])
                        update_lead_voice_extraction(
                            contact_id, tenant_id=tenant_id,
                            customer_name=fields.get("customer_name"),
                            property_type=fields.get("property_type"),
                            budget=fields.get("budget"),
                        )
                    except Exception as exc:
                        logger.warning(
                            "חילוץ שדות אוטומטי מהודעה קולית נכשל (התמלול עצמו הצליח): %s", exc,
                        )
            else:
                message_text = voice_result["text"]

    if not contact_id or not message_text:
        return Response(status=400)

    _log_incoming_message(tenant_id, source, contact_id, message_text)
    db.log_message(contact_id, message_text, direction="in", tenant_id=tenant_id, channel=source)

    reply_text = None
    try:
        if source == "whatsapp":
            # ל-whatsapp מייצרים גם תשובה אוטומטית ושולחים אותה חזרה ללקוח דרך TwiML.
            # זו תגובה בתוך חלון השיחה שהלקוח פתח - לא הודעה יזומה - ולכן אינה דורשת
            # הודעת-תבנית מאושרת מול מטא.
            card, reply_text = process_message_with_reply(
                contact_id, message_text, tenant_id=tenant_id, source_channel=source
            )
            db.log_message(contact_id, reply_text, direction="out", tenant_id=tenant_id, channel=source)
        else:
            card = process_message(contact_id, message_text, tenant_id=tenant_id, source_channel=source)
        logger.info("[tenant=%s] [%s] עודכן כרטיס לקוח: %s", tenant_id, source, card)
    except Exception as exc:
        logger.error("שגיאה בעיבוד הודעה מ-%s (tenant=%s): %s", source, tenant_id, exc, exc_info=True)

    if source == "whatsapp":
        # Twilio מצפה לתגובת TwiML; <Message> בתוכה נשלח כתשובה בוואטסאפ ללקוח
        body = f"<Message>{escape(reply_text)}</Message>" if reply_text else ""
        twiml = f"<?xml version='1.0' encoding='UTF-8'?><Response>{body}</Response>"
        return Response(twiml, mimetype="text/xml")

    return Response(status=200)


if __name__ == "__main__":
    logger.info(
        "מפעיל שרת על %s:%s (debug=%s, verify_twilio_signature=%s, log_level=%s)",
        HOST, PORT, FLASK_DEBUG, VERIFY_TWILIO_SIGNATURE, LOG_LEVEL,
    )
    scheduler.start()
    app.run(host=HOST, port=PORT, debug=FLASK_DEBUG, use_reloader=False)
